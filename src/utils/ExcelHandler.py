from openpyxl import workbook
import openpyxl


def read_excel(filepath,sheetname):
    wb = openpyxl.load_workbook(filepath)
    work_sheet = wb[sheetname]
    max_row   = work_sheet.max_row
    max_column = work_sheet.max_column
    excel_data = []

    for row in work_sheet.values:
        row_list =[]
        for value in row:
            row_list.append(value)
        excel_data.append(row_list)
    return excel_data


def write_excel(filepath,sheetname,list_of_list_data):
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(title=sheetname)

    for i in range(0,len(list_of_list_data)):
        row = list_of_list_data[i]
        for j in range(0,len(row)):
            ws.cell(column=j+1,row=i+1,value=row[j])

    wb.save(filename=filepath)


def writing_resultset(filename,sheetname,resultset):
    col_names = list(map(lambda x:x[0],resultset.description))
    #col_names = [row[0] for row in resultset.description]
    resultset = list(resultset.fetchall())
    resultset.insert(0,col_names)
    write_excel(filename,sheetname,resultset)



